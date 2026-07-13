#!/usr/bin/env python3
"""
05_reading_list.py
==================
Builds the close-reading list from the min-5 cluster typology + the validated
interactional lexicon audit:
  - anchors : most-cited and most-typical paper(s) per cluster (Sec 7 exemplars)
  - deviant : records that use any validated interactional term (RICH vs metric)
  - assignments : every keyword-assignable record -> dominant min-5 cluster
Outputs CSVs to outputs/ and a workbook CRS_reading_list.xlsx.
Deterministic; logs via runlog.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from runlog import log_run


import pandas as pd, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pathlib import Path
BASE=Path("ConnversationalRSs"); OUT=BASE/"outputs"
def newest(patterns, fallback):
    for pat in patterns:
        c=sorted(OUT.glob(pat))
        if c: return c[-1]
    return Path(fallback)
CORPUS=newest(["merged_scopus_wos_*.csv"], "/mnt/project/merged_scopus_wos_20260625.csv")
CTERMS=Path("/mnt/project/clusterterms_Min5_Res0_6_250626.csv")
if (OUT/"clusterterms_Min5_Res0_6_250626.csv").exists(): CTERMS=OUT/"clusterterms_Min5_Res0_6_250626.csv"
AUDIT=newest(["interactional_lexicon_audit_*.csv"], "/mnt/user-data/uploads/interactional_lexicon_audit_2026-06-30.csv")
EXCLUDE={'face','next turn','coherence','overlap','accommodation'}
METRIC={'trust','engagement'}  # present but metric-sense, not genuinely interactional

ROLE={1:"elicitation",3:"elicitation",7:"elicitation",4:"signal",5:"signal",
      2:"input_to_learn",6:"reflexive",8:"reflexive"}
LABEL={1:"C1 Preference elicitation & filtering",2:"C2 Neural/graph/LLM learning",
       3:"C3 Conversational agents & chat UIs",4:"C4 Knowledge-based critiquing",
       5:"C5 Dialogue systems & NLP",6:"C6 Evaluation & user simulation",
       7:"C7 Conversational UI design",8:"C8 Explainability"}

m=pd.read_csv(CORPUS,dtype=str,keep_default_na=False)
m["cit"]=pd.to_numeric(m["Cited by"],errors="coerce").fillna(0).astype(int)
m["first_author"]=m["Authors"].str.split(";").str[0].str.strip()
ct=pd.read_csv(CTERMS)
cterms={c:set(ct[ct.cluster==c].term.str.lower()) for c in ct.cluster.unique()}
tocc={(r.cluster,r.term.lower()):r.occurrences for r in ct.itertuples()}
kws=lambda s:[k.strip().lower() for k in re.split(r"[;,]",s) if k.strip()]

rows=[]
for i,r in m.iterrows():
    ks=kws(r["Author Keywords"])
    if not ks: rows.append((None,0,0.0,"")); continue
    best,bn,bw,bm=None,0,0,[]
    for c,terms in cterms.items():
        mt=[k for k in ks if k in terms]; w=sum(tocc.get((c,k),0) for k in mt)
        if len(mt)>bn or (len(mt)==bn and w>bw): best,bn,bw,bm=c,len(mt),w,mt
    rows.append((best,bn,round(bn/len(ks),2),"; ".join(bm)))
m["cluster"],m["kw_matches"],m["kw_share"],m["matched_kw"]=zip(*rows)
asg=m[m.cluster.notna()].copy()

# anchors: most-cited + most-typical per cluster
anc=[]
for c in sorted(cterms):
    sub=asg[asg.cluster==c]
    if not len(sub): continue
    picks={}
    for _,r in sub.sort_values("cit",ascending=False).head(2).iterrows(): picks[r["Title"]]=("most-cited",r)
    for _,r in sub[sub.kw_matches>=2].sort_values(["kw_matches","cit"],ascending=False).head(2).iterrows():
        picks.setdefault(r["Title"],("most-typical",r))
    for t,(kind,r) in picks.items():
        anc.append({"role":ROLE[c],"cluster":LABEL[c],"pick":kind,"cited_by":r.cit,
                    "year":r["Year"],"first_author":r["first_author"],"title":r["Title"],
                    "matched_keywords":r["matched_kw"],"source":r["Source title"][:45],"DOI":r["DOI"]})
anc=pd.DataFrame(anc).sort_values(["role","cluster","pick"])

# deviant cases
aud=pd.read_csv(AUDIT)
if "excluded_after_inspection" in aud.columns:
    lex=aud[(aud.n_records>0)&(~aud["excluded_after_inspection"])]
else:
    lex=aud[(aud.n_records>0)&(~aud.term.isin(EXCLUDE))]
tbl={l:list(g.term) for l,g in lex.groupby("layer")}
m["text"]=(m["Title"].fillna("")+". "+m["Abstract"].fillna(""))
dev=[]
for _,r in m.iterrows():
    hits=[]
    for layer,terms in tbl.items():
        for t in terms:
            if re.search(r"\b"+re.escape(t)+r"\b",r["text"],re.I): hits.append((t,layer))
    if hits:
        genuine=[t for t,l in hits if t not in METRIC]
        dev.append({"priority":"RICH" if genuine else "metric-only",
                    "cluster":LABEL.get(int(r.cluster),"") if pd.notna(r.cluster) else "(unassigned)",
                    "role":ROLE.get(int(r.cluster),"") if pd.notna(r.cluster) else "",
                    "cited_by":r.cit,"year":r["Year"],"first_author":r["first_author"],
                    "interactional_terms":"; ".join(f"{t}[{l}]" for t,l in hits),
                    "title":r["Title"],"DOI":r["DOI"]})
dev=pd.DataFrame(dev).sort_values(["priority","cited_by"],ascending=[True,False])

# assignments reference
ref=asg[["cluster","kw_matches","kw_share","cit","year","first_author","Title","matched_kw","DOI"]].copy() if False else None
ref=asg.copy(); ref["cluster"]=ref["cluster"].map(LABEL)
ref=ref[["cluster","kw_matches","kw_share","cit","Year","first_author","Title","matched_kw","DOI"]]
ref.columns=["cluster","kw_matches","kw_share","cited_by","year","first_author","title","matched_keywords","DOI"]
ref=ref.sort_values(["cluster","cited_by"],ascending=[True,False])

# ---- write workbook ----
wb=Workbook(); ws=wb.active; ws.title="Read me"
readme=[
 ["CRS reading list — anchors and deviant cases",""],
 ["Generated from the min-5 cluster typology + the validated interactional lexicon audit.",""],
 ["",""],
 ["Sheet","Purpose"],
 ["Anchors","1-2 most-cited and most-typical papers per cluster. Use for the CA mapping in Sec 7:"],
 ["","  one anchor per role (elicitation / signal / input-to-learn) to read closely."],
 ["Deviant cases","The 70 papers that use interactional vocabulary at all. priority=RICH means they use"],
 ["","  a genuinely interactional term (mixed-initiative, rapport, dialogue state tracking, etc.);"],
 ["","  priority=metric-only means only trust/engagement (evaluation constructs, weaker signal)."],
 ["","  These test whether the words reflect real interactional treatment (lexical vs conceptual)."],
 ["Cluster assignments","All 497 keyword-assignable records mapped to their dominant min-5 cluster (reference/sampling)."],
 ["",""],
 ["Notes",""],
 ["","Cluster = dominant min-5 author-keyword cluster; records with no min-5 keyword are unassigned (369)."],
 ["","CA-core layer contributes ZERO deviant cases — the interactional presence is HCI/CUI + 2 pragmatics."],
 ["","'most-typical' = highest count of in-cluster author keywords; 'most-cited' = highest Cited by."],
]
for row in readme: ws.append(row)
ws["A1"].font=Font(bold=True,size=13)
for r in (4,13): 
    ws[f"A{r}"].font=Font(bold=True); ws[f"B{r}"].font=Font(bold=True)
ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=95

def add(df,name):
    ws=wb.create_sheet(name)
    ws.append(list(df.columns))
    for _,r in df.iterrows(): ws.append([r[c] for c in df.columns])
    hdr=PatternFill("solid",start_color="4C78A8")
    for c in range(1,len(df.columns)+1):
        cell=ws.cell(1,c); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=hdr
        cell.alignment=Alignment(vertical="center")
    ws.freeze_panes="A2"
    widths={"title":60,"matched_keywords":34,"interactional_terms":40,"DOI":26,
            "source":34,"first_author":18,"cluster":30,"role":14,"pick":12,"priority":11}
    for i,col in enumerate(df.columns,1):
        ws.column_dimensions[get_column_letter(i)].width=widths.get(col,10)
    return ws

add(anc,"Anchors")
dws=add(dev,"Deviant cases")
rich=PatternFill("solid",start_color="FFF3CD")
for r in range(2,len(dev)+2):
    if dws.cell(r,1).value=="RICH":
        dws.cell(r,1).fill=rich
add(ref,"Cluster assignments")

out=str(OUT/"CRS_reading_list.xlsx")
wb.save(out)
anc.to_csv(OUT/"anchors.csv",index=False)
dev.to_csv(OUT/"deviant_cases.csv",index=False)
ref.to_csv(OUT/"doc_cluster_assignments.csv",index=False)
log_run(__file__, inputs=[str(CORPUS),str(CTERMS),str(AUDIT)],
        outputs=[out, str(OUT/"anchors.csv"), str(OUT/"deviant_cases.csv"),
                 str(OUT/"doc_cluster_assignments.csv")],
        params=f"min5 cluster assignment via author keywords; anchors=most-cited+most-typical; "
               f"deviant=validated interactional lexicon; excluded={sorted(EXCLUDE)}",
        notes=f"anchors={len(anc)}, deviant={len(dev)}, assigned={len(ref)}")
print("anchors:",len(anc),"| deviant:",len(dev),"(RICH:",int((dev.priority=='RICH').sum()),
      "metric-only:",int((dev.priority=='metric-only').sum()),") | assigned:",len(ref))
print("saved workbook + 3 CSVs to outputs")
